from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout, update_session_auth_hash
from .forms import AdminForm, AgentF, AnnouncementForm, BranchF, CeoMessageForm,DepartmentHeadForm, DownloadForm, FiscalForm, OtherDownloadForm, RIForm, SettingForm, SurveryorF, CitizenF, ReportF, TermForm, helpForm, newsF, BodForm, ManagementForm, ProductFrom, SubProductFrom, QuestionAnswerFrom, socialSiteForm
from landing.models import Announcement, CeoMessage, Branch, Contact, DepartmentHead, OtherDownload, PageVisit, RIpartner, Setting, Term, TopBar, fiscalYear, helpCenter, news, Surveryor, Agent, Citizen, Report, Download, Bod, ManagementTeam, Product, Sub_product, QuestionAnswer, socialSite
from django.contrib import messages
from django.http.response import HttpResponseRedirect
from django.urls import reverse
import datetime
from .forms import FormChangePassword
import openpyxl
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.models import User

# Create your views here.


def getOrdering(request):
    Rid = request.GET.get('tourId', None)
    que = Report.objects.filter(rtype = Rid).count() + 1
    
    data = {
            'order':que
        }
    return JsonResponse(data)

@login_required
def dashboard(request):
    news_count = news.objects.all().count()
    branch_count = Branch.objects.all().count()
    agent_count = Agent.objects.all().count()
    citizen_count = Citizen.objects.all().count()
    report_count = Report.objects.all().count()
    survey_count = Surveryor.objects.all().count()
    product_count = Product.objects.all().count()
    sub_product_count = Sub_product.objects.all().count()
    faq_count = QuestionAnswer.objects.all().count()
    form_count = Download.objects.all().count()
    depart_count = DepartmentHead.objects.all().count()
    bod_count = Bod.objects.all().count()
    manage_count = ManagementTeam.objects.all().count()
    page_count = PageVisit.objects.all().count()
    fiscal = fiscalYear.objects.all().count()
    
    
    dist ={
        'news_count':news_count,
        'branch_count':branch_count,
        'agent_count':agent_count,
        'citizen_count': citizen_count,
        'report_count':report_count,
        'survey_count':survey_count,
        'product_count':product_count,
        'sub_product_count':sub_product_count,
        'faq_count':faq_count,
        'form_count':form_count,
        'depart_count':depart_count,
        'bod_count':bod_count,
        'manage_count':manage_count,
        'count':page_count,
        'fiscal':fiscal
        }
    return render(request, "manager/dashboard.html", dist)

def logoutUser(request):
    logout(request)
    return HttpResponseRedirect(reverse('landing:landing'))
    


def loadData(request):
    if request.method == 'POST':
        csvFiles = request.FILES.get('excelfile')
        wb = openpyxl.load_workbook(csvFiles, data_only=True)
        worksheet = wb["Sheet1"]
        print(worksheet)
        # try:
        for row in worksheet.iter_rows(min_row=2):
            print(row)
            full_data = []
            for cell in row:
                full_data.append(cell.value)

            if not full_data[0] == None:
                name = full_data[0]
                address = full_data[1]
                contact = full_data[2]
                lience = full_data[3]
                issue = full_data[4]
                print(issue)
                email = full_data[5]
                agent_code = full_data[6]
                # iss = datetime.datetime.strftime
                Agent.objects.create(ordering = Agent.objects.all().count()+1 ,name=name, agent_code=agent_code, address=address, contact= contact, lience_no= lience,  email=email, issue_date= issue)
        messages.success(request, "Successfully Uploaded From excel File")
        # except:
        #     messages.success(request, "No data to Load. Something went wrong")
        return HttpResponseRedirect(reverse('manager:addAgent'))



def loadDataSuv(request):
    if request.method == 'POST':
        csvFiles = request.FILES.get('excelfile')
        wb = openpyxl.load_workbook(csvFiles, data_only=True)
        print(wb)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        # try:
        for row in worksheet.iter_rows(min_row=2):
            print(row)
            full_data = []
            for cell in row:
                full_data.append(cell.value)
    
            if not full_data[0] == None:
                name = full_data[0]
                specilization = full_data[1]
                lience = full_data[2]
                issue = full_data[3]
                renew = full_data[4]
                area = full_data[5]
                contact = full_data[6]
                email = full_data[7]
                Surveryor.objects.create(ordering = Surveryor.objects.all().count()+1, name=name, specilization=specilization, contact= contact, lience_no= lience,  email=email, issued_date= issue, area=area, renew_date=renew)
        messages.success(request, "Successfully Uploaded From excel File")
        # except:
        #     messages.success(request, "No data to Load. Something went wrong")
        return HttpResponseRedirect(reverse('manager:addSurveryor'))


def addAgent(request):
    form = AgentF(request.POST or None)
    agent = Agent.objects.all().order_by('ordering')
    form.fields['ordering'].initial = agent.count()+1
    page = request.GET.get('page', 1)

    paginator = Paginator(agent, 40)
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    dist = {
        'form':form,
        'data':data
    }
    if form.is_valid():
        form.save()
        messages.success(request, "Agents Saved Succesfully")
        return HttpResponseRedirect(reverse('manager:addAgent'))
    else:
        return render(request, 'manager/addAgent.html',dist )
       
    return render(request, 'manager/addAgent.html',dist )


def addBranch(request):
    form = BranchF()
    branched = Branch.objects.all().order_by('-id')
    page = request.GET.get('page', 1)

    paginator = Paginator(branched, 40)
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    dist = {
        'form':form,
        'branch':data,
    }
    if request.method == 'POST':
        form = BranchF(request.POST, request.FILES or None)
        if form.is_valid():        
            form.save()
            messages.success(request, "Branch Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:addBranch'))
        else:
            return render(request, 'manager/addBranch.html',dist )
    else:
        return render(request, 'manager/addBranch.html',dist )

def deleteBranch(request, id):
    branch = Branch.objects.get(id = id)
    branch.delete()
    messages.success(request, "Successfully Deleted Branch")
    return HttpResponseRedirect(reverse('manager:addBranch'))


def addSurveryor(request):
    form = SurveryorF(request.POST or None)
    agent = Surveryor.objects.all().order_by('ordering')
    form.fields['ordering'].initial = agent.count()+1
    page = request.GET.get('page', 1)

    paginator = Paginator(agent, 40)
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    dist = {
        'form':form,
        'data':data
    }
    if form.is_valid():
        form.save()
        messages.success(request, "Surveryor Saved Succesfully")
        return HttpResponseRedirect(reverse('manager:addSurveryor'))
    else:
        return render(request, 'manager/addSurveryor.html',dist )
       
    return render(request, 'manager/addSurveryor.html',dist )


def addCitizen(request):
    form = CitizenF(request.POST or None)
    agent = Citizen.objects.all().order_by('ordering')
    form.fields['ordering'].initial = agent.count()+1
    dist = {
        'form':form,
        'data':agent
    }
    if form.is_valid():
        form.save()
        messages.success(request, "Citizen Saved Succesfully")
        return HttpResponseRedirect(reverse('manager:addCitizen'))
    else:
        return render(request, 'manager/addCitizen.html',dist )
       
    return render(request, 'manager/addCitizen.html',dist )

def addNews(request):
    form = newsF()
    dist = {
        'form':form
    }
    if request.method == 'POST':
        forms = newsF(request.POST, request.FILES)
        if forms.is_valid():
            forms.save()
            messages.success(request, "News Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:addNews'))
        else:
            messages.error(request, "Something went Wrong")
            return HttpResponseRedirect(reverse('manager:addNews'))
    else:
        return render(request, 'manager/addNews.html',dist )



def editNews(request, id):
    form = newsF()
    new = news.objects.get(id = id)
    form.fields['name'].initial = new.name
    form.fields['description'].initial = new.description
    form.fields['dateof'].initial = new.dateof

    dist = {
        'form':form,
        'news':new
    }
    if request.method == 'POST':
        forms = newsF(request.POST, request.FILES)
        if forms.is_valid():
            cd = forms.cleaned_data
            new.name = cd['name']
            new.description = cd['description']
            new.dateof = cd['dateof']
            if cd['files'] != None:
                new.files = cd['files']
            if cd['image'] != None:
                new.image = cd['image']
            new.save()
            messages.success(request, "News Updated Succesfully")
            return HttpResponseRedirect(reverse('manager:editNews',args=[new.id]))
        else:
            messages.error(request, "Something went Wrong")
            return HttpResponseRedirect(reverse('manager:editNews'))
    else:
        return render(request, 'manager/editNews.html',dist )



def addReport(request):
    form = ReportF()
    form.fields['ordering'].initial = Report.objects.all().count()+1
    dist = {
        'form':form
    }
    if request.method == 'POST':
        forms = ReportF(request.POST, request.FILES)
        if forms.is_valid():
            forms.save()
            messages.success(request, "Report Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:addReport'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/addReport.html',dist )
    else:
        return render(request, 'manager/addReport.html',dist )

def editReport(request, id):
    form = ReportF()
    bod = Report.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['rtype'].initial = bod.rtype
    form.fields['fiscal'].initial = bod.fiscal
    form.fields['ordering'].initial = bod.ordering

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.rtype = request.POST['rtype']
        bod.fiscal = fiscalYear.objects.get(id = request.POST['fiscal'])
        bod.ordering = request.POST['ordering']
        if request.FILES:
            bod.files = request.FILES['files']
        bod.save()
        messages.success(request, "Report Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editReport', args=[bod.id]))
    else:
        return render(request, 'manager/editReport.html',dist )


def editAnnouncement(request, id):
    form = AnnouncementForm()
    bod = Announcement.objects.get(id = id)
    form.fields['ordering'].initial = bod.ordering
    form.fields['name'].initial = bod.name
    form.fields['description'].initial = bod.description
    form.fields['link'].initial = bod.link

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.ordering = request.POST['ordering']
        bod.name = request.POST['name']
        bod.description = request.POST['description']
        bod.link = request.POST['link']
        if request.FILES:
            bod.image = request.FILES['image']
        bod.save()
        messages.success(request, "Announcement Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editAnnoucement', args=[bod.id]))
    else:
        return render(request, 'manager/editAnnoucement.html', dist )
def formReport(request, id):
    if id == 1:
        report = Download.objects.filter(dtype = 'KYC Form').order_by("-id")
        name = "KYC Form"
    elif id == 2:
        report = Download.objects.filter(dtype = 'Proposal Form').order_by("-id")
        name = "Proposal Form"
    else:
        report = Download.objects.filter(dtype = 'Claim Form').order_by("-id")
        name = "Claim Form"

    dist ={
        'name':name,
        'report':report,
        'type':id
    }
    return render(request, 'manager/viewForm.html', dist)

def statementReport(request, id):
    if id == 1:
        report = Report.objects.filter(rtype = 'Annual Report').order_by("ordering")
        name = "Annual"
    elif id == 2:
        report = Report.objects.filter(rtype = 'Quarterly Report').order_by("fiscal")
        name = "Quarterly"
    elif id == 3:
        report = Report.objects.filter(rtype = 'Minute Report').order_by("ordering")
        name = "Minute"
    else:
        report = Report.objects.filter(rtype = 'Other Report').order_by("ordering")
        name = "Other"
    if request.method == 'POST':
        search = request.POST['search']
        report = report.filter(name__icontains= search)
    dist ={
        'name':name,
        'report':report,
        'type':id
    }
    return render(request, 'manager/viewReport.html', dist)

def newUpdate(request):
    report = news.objects.all().order_by("-dateof")
    name = "News and Update"
    page = request.GET.get('page', 1)

    paginator = Paginator(report, 40)
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    if request.method == 'POST':
        search = request.POST['search']
        report = news.objects.filter(name__icontains= search)
    dist ={
        'name':name,
        'report':data
    }
    return render(request, 'manager/news.html', dist)

def setting(request):
    sett = Setting.objects.all()
    form = SettingForm()
    setting = None
    if sett:
        for i in sett:
            setting = i 
            break
        form.fields['email'].initial = setting.email
        form.fields['product_email'].initial = setting.product_email
        form.fields['number1'].initial = setting.number1
        form.fields['number2'].initial = setting.number2
        form.fields['location'].initial = setting.location
        form.fields['toll_free_no'].initial = setting.toll_free_no

    dist = {
        'setting':setting,
        'form':form
    }
    if request.method == 'POST':
        form = SettingForm(request.POST, request.FILES)
        if form.is_valid():
            
            if sett:
                setting.email = request.POST['email']
                setting.number1 = request.POST['number1']
                setting.number2 = request.POST['number2']
                setting.product_email = request.POST['product_email']
                setting.location = request.POST['location']
                setting.toll_free_no = request.POST['toll_free_no']
                try:
                   setting.logo = request.FILES['logo']
                except:
                    pass
                setting.save()
            else:
                form.save()
            messages.success(request, "Successfully Updated Settings")
            return HttpResponseRedirect(reverse("manager:setting"))
    else:
        return render(request, "manager/settings.html", dist)


def addBod(request):
    form = BodForm()
    form.fields['ordering'].initial = Bod.objects.all().count() + 1
    branched = Bod.objects.all().order_by('ordering')
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        form = BodForm(request.POST, request.FILES or None)
        if form.is_valid():        
            form.save()
            messages.success(request, "Bod Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:addBod'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/addBod.html',dist )
    else:
        return render(request, 'manager/addBod.html',dist )

def managementTeam(request):
    form = ManagementForm()
    branched = ManagementTeam.objects.all().order_by('ordering')
    form.fields['ordering'].initial = ManagementTeam.objects.all().count() + 1
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        form = ManagementForm(request.POST, request.FILES or None)
        if form.is_valid():        
            form.save()
            messages.success(request, "Management Team Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:managementTeam'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/managementTean.html',dist )
    else:
        return render(request, 'manager/managementTean.html',dist )


def ProductView(request):
    form = ProductFrom()
    branched = Product.objects.all().order_by('ordering')
    form.fields['ordering'].initial = Product.objects.all().count() + 1
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        form = ProductFrom(request.POST, request.FILES or None)
        if form.is_valid():        
            form.save()
            messages.success(request, "Product Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:product'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/product.html',dist )
    else:
        return render(request, 'manager/product.html',dist )
    

def SubProductView(request):
    form = SubProductFrom()
    branched = Sub_product.objects.all().order_by('ordering')
    form.fields['ordering'].initial = branched.count()+1
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        form = SubProductFrom(request.POST, request.FILES or None)
        if form.is_valid():        
            form.save()
            messages.success(request, "Sub Product Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:subProduct'))
        else:
            messages.success(request, "Something went Wrong")
            return HttpResponseRedirect(reverse('manager:subProduct'))
    else:
        return render(request, 'manager/sub_product.html',dist )


def QuestionAnswerView(request):
    form = QuestionAnswerFrom()
    branched = QuestionAnswer.objects.all().order_by('ordering')
    form.fields['ordering'].initial = QuestionAnswer.objects.all().count() + 1
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        form = QuestionAnswerFrom(request.POST)
        if form.is_valid():        
            form.save()
            messages.success(request, "Question Answer Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:questionAnswer'))
        else:
            messages.success(request, "Something went Wrong")
            return HttpResponseRedirect(reverse('manager:questionAnswer'))
    else:
        return render(request, 'manager/questionAnswer.html',dist )

def AnnouncementView(request):
    form = AnnouncementForm()
    branched = Announcement.objects.all().order_by('-id')
    form.fields['ordering'].initial = branched.count()+1
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, request.FILES)
        if form.is_valid():        
            form.save()
            messages.success(request, "Announcement Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:announcement'))
        else:
            messages.success(request, "Something went Wrong")
            return HttpResponseRedirect(reverse('manager:announcement'))
    else:
        return render(request, 'manager/announcement.html',dist )


def addForm(request):
    form = DownloadForm()

    dist = {
        'form':form,
    }
    if request.method == 'POST':
        forms = DownloadForm(request.POST, request.FILES)
        if forms.is_valid():        
            forms.save()
            messages.success(request, "Forms Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:addForm'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/addForm.html',dist )
    else:
        return render(request, 'manager/addForm.html',dist)


def DepartmentView(request):
    form = DepartmentHeadForm()
    branched = DepartmentHead.objects.all().order_by('ordering')
    form.fields['ordering'].initial = DepartmentHead.objects.all().count() + 1
    dist = {
        'form':form,
        'bod':branched,
    }
    if request.method == 'POST':
        forms = DepartmentHeadForm(request.POST, request.FILES or None)
        if forms.is_valid():        
            forms.save()
            messages.success(request, "Department Team Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:departmentTeam'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/departmentTeam.html',dist )
    else:
        return render(request, 'manager/departmentTeam.html',dist )


def fiscalView(request):
    form = FiscalForm()
    branched = fiscalYear.objects.all().order_by('-id')
    dist = {
        'form':form,
        'fiscal':branched,
    }
    if request.method == 'POST':
        forms = FiscalForm(request.POST)
        for i in branched:
            if i.fiscal == request.POST['fiscal']:
                messages.success(request, "Fiscal Year has been added Already")
                return HttpResponseRedirect(reverse('manager:addFiscal'))
            else:
                pass
        if forms.is_valid():
            forms.save()
            messages.success(request, "Fiscal Year Saved Succesfully")
            return HttpResponseRedirect(reverse('manager:addFiscal'))
        else:
            messages.success(request, "Something went Wrong")
            return render(request, 'manager/addFiscalYear.html',dist )
    else:
        return render(request, 'manager/addFiscalYear.html',dist )


def deleteFiscal(request, id):
    aa = fiscalYear.objects.get(id = id)
    aa.delete()
    messages.success(request, "Sucessfully Delete Fiscal Year")

    return HttpResponseRedirect(reverse('manager:addFiscal'))

def deleteAnn(request, id):
    aa=  Announcement.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Announcement")
    return HttpResponseRedirect(reverse('manager:announcement'))

def deleteProduct(request, id):
    aa=  Product.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Product")
    return HttpResponseRedirect(reverse('manager:product'))

def deleteSubProduct(request, id):
    aa =  Sub_product.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Sub-Product")
    return HttpResponseRedirect(reverse('manager:subProduct'))

def deleteQuestion(request, id):
    aa =  QuestionAnswer.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Questions")
    return HttpResponseRedirect(reverse('manager:questionAnswer'))

def deleteDepartment(request, id):
    aa =  DepartmentHead.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Department Team")
    return HttpResponseRedirect(reverse('manager:departmentTeam'))


def deleteAgent(request, id):
    aa =  Agent.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Agent")
    return HttpResponseRedirect(reverse('manager:addAgent'))

def deleteCitizen(request, id):
    aa =  Citizen.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Citizen")
    return HttpResponseRedirect(reverse('manager:addCitizen'))

def deleteSuveryor(request, id):
    aa =  Surveryor.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Surveryor")
    return HttpResponseRedirect(reverse('manager:addSurveryor'))

def deleteBod(request, id):
    aa =  Bod.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Bod")
    return HttpResponseRedirect(reverse('manager:addBod'))

def deleteManagement(request, id):
    aa =  ManagementTeam.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Management Team")
    return HttpResponseRedirect(reverse('manager:managementTeam'))

def deleteReport(request, id, type):
    aa =  Report.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Report")
    return HttpResponseRedirect(reverse('manager:statementReport', args=[type]))

def deleteForm(request, id, type):
    aa =  Download.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted Forms")
    return HttpResponseRedirect(reverse('manager:formReport', args=[type]))

def deleteNews(request, id):
    aa =  news.objects.get(id =id )
    aa.delete()
    messages.success(request,"Sucessfully Deleted News")
    return HttpResponseRedirect(reverse('manager:newUpdate'))

def editProduct(request, id):
    form = ProductFrom()
    bod = Product.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['description'].initial = bod.description
    form.fields['discontinue'].initial = bod.discontinue
    form.fields['icons'].initial = bod.icons
    form.fields['ordering'].initial = bod.ordering

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.description = request.POST['description']
        bod.ordering = request.POST['ordering']
        bod.icons = request.POST['icons']
        continues = request.POST.get('discontinue', False)
        if continues == 'on':
            bod.discontinue = True
        else:
            bod.discontinue = False

        try:
            bod.image = request.FILES['image']
        except:
            pass
            
        bod.save()
        messages.success(request, "Product Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editProduct', args=[bod.id]))
    else:
        return render(request, 'manager/editProduct.html',dist )


def editSubProduct(request, id):
    form = SubProductFrom()
    bod = Sub_product.objects.get(id = id)
    form.fields['product'].initial = bod.product
    form.fields['name'].initial = bod.name
    form.fields['description'].initial = bod.description
    form.fields['discontinue'].initial = bod.discontinue
    form.fields['icons'].initial = bod.icons
    form.fields['ordering'].initial = bod.ordering

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.description = request.POST['description']
        bod.icons = request.POST['icons']
        bod.ordering = request.POST['ordering']
        bod.product = Product.objects.get(id = request.POST['product']) 
        continues = request.POST.get('discontinue', False)
        if continues == 'on':
            bod.discontinue = True
        else:
            bod.discontinue = False

        try:
            bod.image = request.FILES['image']
        except:
            pass
            
        bod.save()
        messages.success(request, "Sub Product Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editSubProduct', args=[bod.id]))
    else:
        return render(request, 'manager/editSubProduct.html',dist )

def editQuestion(request, id):
    form = QuestionAnswerFrom()
    bod = QuestionAnswer.objects.get(id = id)
    form.fields['question'].initial = bod.question
    form.fields['answer'].initial = bod.answer
    form.fields['ordering'].initial = bod.ordering
    
    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.question = request.POST['question']
        bod.answer = request.POST['answer']
        bod.ordering = request.POST['ordering']
        bod.save()
        messages.success(request, "FAQ Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editQuestionAns', args=[bod.id]))
    else:
        return render(request, 'manager/editQuestionAns.html',dist )

def editDepartment(request, id):
    form = DepartmentHeadForm()
    bod = DepartmentHead.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['post'].initial = bod.post
    form.fields['email'].initial = bod.email
    form.fields['ordering'].initial = bod.ordering

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.post = request.POST['post']
        bod.ordering = request.POST['ordering']
        bod.email = request.POST['email']
        try:
            bod.image = request.FILES['image']
        except:
            pass
            
        bod.save()
        messages.success(request, "Department Team Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editDepartment', args=[bod.id]))
    else:
        return render(request, 'manager/editDepartment.html',dist )

def editBranch(request, id):
    form = BranchF()
    bod = Branch.objects.get(id = id)
    form.fields['BranchName'].initial = bod.BranchName
    form.fields['district'].initial = bod.district
    form.fields['street'].initial = bod.street
    form.fields['contact'].initial = bod.contact
    form.fields['fax'].initial = bod.fax
    form.fields['email'].initial = bod.email
    form.fields['focal'].initial = bod.focal
    form.fields['provience'].initial = bod.provience
    form.fields['number'].initial = bod.number
    form.fields['focal_img'].initial = bod.focal_img.url

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.BranchName = request.POST['BranchName'] 
        bod.district = request.POST['district']
        bod.street = request.POST['street']
        bod.contact = request.POST['contact']
        bod.fax = request.POST['fax']
        bod.email = request.POST['email']
        bod.focal = request.POST['focal']
        bod.provience = request.POST['provience']
        bod.number = request.POST['number']
    
        try:
            bod.focal_img = request.FILES['focal_img']
        except:
            pass
            
        bod.save()
        messages.success(request, "Branch Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editBranch', args=[bod.id]))
    else:
        return render(request, 'manager/editBranch.html',dist )

def editAgent(request, id):
    form = AgentF()
    bod = Agent.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['address'].initial = bod.address
    form.fields['contact'].initial = bod.contact
    form.fields['email'].initial = bod.email
    form.fields['lience_no'].initial = bod.lience_no
    form.fields['issue_date'].initial = bod.issue_date
    form.fields['agent_code'].initial = bod.agent_code
    form.fields['ordering'].initial = bod.ordering

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.address = request.POST['address']
        bod.contact = request.POST['contact']
        bod.agent_code = request.POST['agent_code']
        bod.email = request.POST['email']
        bod.lience_no = request.POST['lience_no']
        if  request.POST['issue_date']:
            bod.issue_date = request.POST['issue_date']
        bod.ordering = request.POST['ordering']
      
        bod.save()
        messages.success(request, "Agent Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editAgent', args=[bod.id]))
    else:
        return render(request, 'manager/editAgent.html',dist )


def editForm(request, id):
    form = DownloadForm()
    bod = Download.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['dtype'].initial = bod.dtype

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.dtype = request.POST['dtype']
        if request.FILES:
            bod.files = request.FILES['files']
    
        bod.save()
        messages.success(request, "Form Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editForm', args=[bod.id]))
    else:
        return render(request, 'manager/editForm.html',dist )

def editCitizen(request, id):
    form = CitizenF()
    bod = Citizen.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['details'].initial = bod.details
    form.fields['ordering'].initial = bod.ordering

    dist = {
        'form':form,
        'bod':bod,
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.details = request.POST['details']
        bod.ordering = request.POST['ordering']
    
        bod.save()
        messages.success(request, "Citizen Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editCitizen', args=[bod.id]))
    else:
        return render(request, 'manager/editCitizen.html',dist )

def editSurvey(request, id):
    form = SurveryorF()
    bod = Surveryor.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['specilization'].initial = bod.specilization
    form.fields['area'].initial = bod.area
    form.fields['contact'].initial = bod.contact
    form.fields['email'].initial = bod.email
    form.fields['lience_no'].initial = bod.lience_no
    form.fields['issued_date'].initial = bod.issued_date
    form.fields['renew_date'].initial = bod.renew_date
    form.fields['ordering'].initial = bod.ordering
    issue = bod.issued_date
    renew = bod.renew_date
    if issue: 
        start_date = datetime.datetime.strptime(str(issue), "%Y-%m-%d").date()
    else:
        start_date = None
    if renew:
        end_date = datetime.datetime.strptime(str(renew), "%Y-%m-%d").date()
    else:
        end_date = None
    print(start_date)
    print(end_date)
    dist = {
        'form':form,
        'bod':bod,
        'issue':str(start_date),
        'renew':str(end_date),
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.area = request.POST['area']
        bod.specilization = request.POST['specilization']
        bod.contact = request.POST['contact']
        bod.email = request.POST['email']
        bod.ordering = request.POST['ordering']
        bod.lience_no = request.POST['lience_no']
        first_date = request.POST['issued_date']
        last_date = request.POST['renew_date']
        if first_date:
            start_date = datetime.datetime.strptime(first_date, "%Y-%m-%d").date()
            bod.issued_date =  first_date
        if last_date:
            end_date = datetime.datetime.strptime(last_date, "%Y-%m-%d").date()
            bod.renew_date = last_date
    
        bod.save()
        messages.success(request, "Surveryor Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editSurvey', args=[bod.id]))
    else:
        return render(request, 'manager/editSurvery.html',dist )


def editBod(request, id):
    form = BodForm()
    bod = Bod.objects.get(id = id)
    form.fields['ordering'].initial = bod.ordering
    form.fields['name'].initial = bod.name
    form.fields['post'].initial = bod.post
    form.fields['type'].initial = bod.type
    form.fields['email'].initial = bod.email
    form.fields['phone'].initial = bod.phone
    form.fields['description'].initial = bod.description
    form.fields['chairman'].initial = bod.chairman
    issue = bod.appointed_date
    renew = bod.re_appointed_date
    if issue: 
        start_date = datetime.datetime.strptime(str(issue), "%Y-%m-%d").date()
    else:
        start_date = None
    if renew:
        end_date = datetime.datetime.strptime(str(renew), "%Y-%m-%d").date()
    else:
        end_date = None
    print(start_date)
    print(end_date)
    dist = {
        'form':form,
        'bod':bod,
        'issue':str(start_date),
        'renew':str(end_date),
    }
    if request.method == 'POST':
        bod.ordering = request.POST['ordering']
        bod.name = request.POST['name']
        bod.post = request.POST['post']
        bod.type = request.POST['type']
        bod.email = request.POST['email']
        bod.phone = request.POST['phone']
        bod.description = request.POST['description']
        continues = request.POST.get('chairman', False)
        if continues == 'on':
            bod.chairman = True
        else:
            bod.chairman = False

        try:
            bod.image = request.FILES['image']
        except:
            pass
        first_date = request.POST['appointed_date']
        last_date = request.POST['re_appointed_date']
        if first_date:
            start_date = datetime.datetime.strptime(first_date, "%Y-%m-%d").date()
            bod.appointed_date =  first_date
        if last_date:
            end_date = datetime.datetime.strptime(last_date, "%Y-%m-%d").date()
            bod.re_appointed_date = last_date
    
        bod.save()
        messages.success(request, "Bod Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editBod', args=[bod.id]))
    else:
        return render(request, 'manager/editBod.html',dist )


def editManagement(request, id):
    form =ManagementForm()
    bod = ManagementTeam.objects.get(id = id)
    form.fields['name'].initial = bod.name
    form.fields['post'].initial = bod.post
    form.fields['email'].initial = bod.email
    form.fields['phone'].initial = bod.phone
    form.fields['ordering'].initial = bod.ordering
    
    issue = bod.appointed_date
    renew = bod.re_appointed_date
    if issue: 
        start_date = datetime.datetime.strptime(str(issue), "%Y-%m-%d").date()
    else:
        start_date = None
    if renew:
        end_date = datetime.datetime.strptime(str(renew), "%Y-%m-%d").date()
    else:
        end_date = None
    print(start_date)
    print(end_date)
    dist = {
        'form':form,
        'bod':bod,
        'issue':str(start_date),
        'renew':str(end_date),
    }
    if request.method == 'POST':
        bod.name = request.POST['name']
        bod.post = request.POST['post']
        bod.email = request.POST['email']
        bod.phone = request.POST['phone']
        bod.ordering = request.POST['ordering']

        try:
            bod.image = request.FILES['image']
        except:
            pass
        first_date = request.POST['appointed_date']
        last_date = request.POST['re_appointed_date']
        if first_date:
            start_date = datetime.datetime.strptime(first_date, "%Y-%m-%d").date()
            bod.appointed_date =  first_date
        if last_date:
            end_date = datetime.datetime.strptime(last_date, "%Y-%m-%d").date()
            bod.re_appointed_date = last_date
    
        bod.save()
        messages.success(request, "Surveryor Edited Succesfully")
        return HttpResponseRedirect(reverse('manager:editManagement', args=[bod.id]))
    else:
        return render(request, 'manager/editManagement.html',dist )

def changePassword(request):
    if request.method == 'POST':
        form = FormChangePassword(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('manager:changePassword')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = FormChangePassword(request.user)
    return render(request, 'manager/changePassword.html', {
        'form': form,
    })

def hideProduct(request, id):
    pro = Product.objects.get(id = id)
    print(pro.hide)
    if pro.hide: 
        pro.hide = False
        pro.save()
        messages.success(request, "Successfully Unhide the Product")
    else:
        pro.hide = True
        pro.save()
        messages.success(request, "Successfully Hidden the Product")

    return HttpResponseRedirect(reverse('manager:product'))

def hideSubProduct(request, id):
    pro = Sub_product.objects.get(id = id)
    print(pro.hide)
    if pro.hide: 
        pro.hide = False
        pro.save()
        messages.success(request, "Successfully Unhide the Sub Product")
    else:
        pro.hide = True
        pro.save()
        messages.success(request, "Successfully Hidden the Sub Product")

    return HttpResponseRedirect(reverse('manager:subProduct'))

def contact(request):
    cont = Contact.objects.all().order_by('-id')
    dist = {
        'contact': cont
    }
    return render(request, "manager/contact.html", dist)


def otherDownload(request):
    form = OtherDownloadForm()
    files = OtherDownload.objects.all().order_by('-id')
    dist = {
        'form':form,
        'site':files
    }
    if request.method == 'POST':
        form = OtherDownloadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Uploaded Files")
            return HttpResponseRedirect(reverse('manager:otherDownload'))
        else:
            messages.error(request, "Sorry your file Could not be uploaded")
            return HttpResponseRedirect(reverse('manager:otherDownload'))
    return render(request, 'manager/addOther.html', dist)

def deleteOther(request, id):
    it = OtherDownload.objects.get(id = id)
    it.delete()
    messages.success(request,"Successfully deleted files")
    return HttpResponseRedirect(reverse('manager:otherDownload'))

def editOther(request, id):
    site = OtherDownload.objects.get(id = id)
    form = OtherDownloadForm()
    form.fields['name'].initial = site.name
    dist = {
        'form':form,
        'site':site
    }
    if request.method == 'POST':
        site.name = request.POST['name']
        if request.FILES:
            site.files = request.FILES['files']
        site.save()
        messages.success(request,"Successfully Updated Files")
        HttpResponseRedirect(reverse('manager:editOther', args=[site.id]))
    return render(request, "manager/editOther.html", dist)
def addCeoMessage(request):
    form = CeoMessageForm()
    message = CeoMessage.objects.all()
    
    if not message:
        dist = {
            'form':form
        }
        if request.method == 'POST':
            fom = CeoMessageForm(request.POST, request.FILES)
            if fom.is_valid():
                fom.save()
                messages.success(request, "Successfully Added Ceo Message") 
                return HttpResponseRedirect(reverse('manager:addCeoMessage'))
        return render(request, 'manager/addCeoMessage.html', dist)
    else:
        for i in message:
            msg = i
            break
        form.fields['message'].initial = msg.message
        dist = {
            'form':form
        }
        if request.method == 'POST':
            msg.message = request.POST['message']
            msg.save()

            messages.success(request, "Successfully Updated Ceo Message") 
            return HttpResponseRedirect(reverse('manager:addCeoMessage'))
        return render(request, 'manager/addCeoMessage.html', dist)

def addRIpartner(request):
    form = RIForm()
    message = RIpartner.objects.all()
    
    if not message:
        dist = {
            'form':form
        }
        if request.method == 'POST':
            fom = RIForm(request.POST, request.FILES)
            if fom.is_valid():
                fom.save()
                messages.success(request, "Successfully Added RI Partner Description") 
                return HttpResponseRedirect(reverse('manager:addRIpartner'))
        return render(request, 'manager/addPartner.html', dist)
    else:
        for i in message:
            msg = i
            break
        form.fields['description'].initial = msg.description
        dist = {
            'form':form
        }
        if request.method == 'POST':
            msg.description = request.POST['description']
            msg.save()
            messages.success(request, "Successfully Updated RI Partner") 
            return HttpResponseRedirect(reverse('manager:addRIpartner'))
        return render(request, 'manager/addPartner.html', dist)

def term(request):
    form = TermForm()
    message = Term.objects.all()
    
    if not message:
        dist = {
            'form':form
        }
        if request.method == 'POST':
            fom = TermForm(request.POST, request.FILES)
            if fom.is_valid():
                fom.save()
                messages.success(request, "Successfully Added Terms and Conditions") 
                return HttpResponseRedirect(reverse('manager:addTerm'))
        return render(request, 'manager/addTerm.html', dist)
    else:
        for i in message:
            msg = i
            break
        form.fields['description'].initial = msg.description
        dist = {
            'form':form
        }
        if request.method == 'POST':
            msg.description = request.POST['description']
            msg.save()
            messages.success(request, "Successfully Updated Term and Condition") 
            return HttpResponseRedirect(reverse('manager:addTerm'))
        return render(request, 'manager/addTerm.html', dist)

def addSocialSite(request):
    form = socialSiteForm()
    site = socialSite.objects.all().order_by('-id')

    dist = {
        'form':form,
        'site':site
    }

    if request.method == 'POST':
        form = socialSiteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Successfully Saved Social Site")
        else:
            messages.error(request, "Failed to add Social Site")
        return HttpResponseRedirect(reverse('manager:addSocial'))
    return render(request, "manager/addSocial.html", dist)

def editSocial(request, id):
    site = socialSite.objects.get(id = id)
    form = socialSiteForm()
    form.fields['name'].initial = site.name
    form.fields['link'].initial = site.link
    form.fields['icon'].initial = site.icon
    dist ={
        'form':form,
        'site':site
    }
    if request.method == 'POST':
        site.name = request.POST['name']
        site.link = request.POST['link']
        site.icon = request.POST['icon']
        site.save()
        messages.success(request,"Successfully Updated Social Site")
        return HttpResponseRedirect(reverse('manager:editSocial', args=[site.id]))
    return render(request, "manager/editSocial.html", dist)

def deleteSocial(request, id):
    site = socialSite.objects.get(id = id)
    site.delete()
    messages.success(request, "Sucessfully delete site")
    return HttpResponseRedirect(reverse('manager:addSocial'))


def addHelp(request):
    form = helpForm()
    site = helpCenter.objects.all().order_by('-id')

    dist = {
        'form':form,
        'data':site
    }

    if request.method == 'POST':
        form = helpForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Successfully Saved Item")
        else:
            messages.error(request, "Failed to add Item")
        return HttpResponseRedirect(reverse('manager:addHelp'))
    return render(request, "manager/help.html", dist)

def editHelp(request, id):
    site = helpCenter.objects.get(id = id)
    form = helpForm()
    form.fields['title'].initial = site.title
    form.fields['description'].initial = site.description

    dist ={
        'form':form,
        'data':site
    }
    if request.method == 'POST':
        site.title = request.POST['title']
        site.description = request.POST['description']
        site.save()
        messages.success(request,"Successfully Updated Item")
        return HttpResponseRedirect(reverse('manager:editHelp', args=[site.id]))
    return render(request, "manager/editHelp.html", dist)

def deleteHelp(request, id):
    site = helpCenter.objects.get(id = id)
    site.delete()
    messages.success(request, "Sucessfully deleted Item")
    return HttpResponseRedirect(reverse('manager:addHelp'))


def topBar(request):
    top = TopBar.objects.all()
    dist = {
        'data':top
    }
    return render(request, "manager/topbar.html", dist)

def hideBar(request, id):
    site = TopBar.objects.get(id = id)
    if site.hide:
        site.hide = False
    else:
        site.hide = True
    site.save()
    messages.success(request, "Successfully Changed Status of the Bar")
    return HttpResponseRedirect(reverse('manager:topBar'))
    
def addAdmin(request):
    form = AdminForm(request.POST or None)
    use = User.objects.all().order_by('-id')
    dist = {
        'form':form,
        'data':use
    }
    if request.method == 'POST':
        if form.is_valid():
            cd = form.cleaned_data
            email = cd['email']
            password1 = cd['password1']
            User.objects.create_user(email = email, username= email, password = password1)
            messages.success(request,"Successfully Added new Admin")
            return HttpResponseRedirect(reverse('manager:addAdmin'))
        else:
            messages.error(request,"Pease Check the Error Below")
            return render(request, "manager/addAdmin.html", dist)
    return render(request, "manager/addAdmin.html", dist)

def deleteAdmin(request, id):
    use = User.objects.get(id = id)
    use.delete()
    messages.success(request, "Successfully Deleted Admin")
    return HttpResponseRedirect(reverse('manager:addAdmin'))

def deleteReport(request, id, slug):
    if slug == 'report':
        Report.objects.get(id =id ).files.delete(save=True)
        messages.success(request, "Successfully Cleared File")
        return HttpResponseRedirect(reverse('manager:editReport', args=[id]))
    elif slug == 'news':
        news.objects.get(id =id ).files.delete(save=True)
        messages.success(request, "Successfully Cleared File")
        return HttpResponseRedirect(reverse('manager:editNews', args=[id]))
    elif slug == "form":
        Download.objects.get(id = id).files.delete(save=True)
        messages.success(request, "Successfully Cleared File")
        return HttpResponseRedirect(reverse('manager:editForm', args=[id]))


def deleteImage(request, id, slug):
    if slug == 'news':
        news.objects.get(id =id ).image.delete(save = True)
        messages.success(request, "Successfully Cleared Image")
        return HttpResponseRedirect(reverse('manager:editNews', args=[id]))
    elif slug=="department":
        DepartmentHead.objects.get(id =id ).image.delete(save = True)
        messages.success(request, "Successfully Cleared Image")
        return HttpResponseRedirect(reverse('manager:editDepartment', args=[id]))
    elif slug=="bod":
        Bod.objects.get(id =id ).image.delete(save = True)
        messages.success(request, "Successfully Cleared Image")
        return HttpResponseRedirect(reverse('manager:editBod', args=[id]))
    elif slug=="management":
        ManagementTeam.objects.get(id =id ).image.delete(save = True)
        messages.success(request, "Successfully Cleared Image")
        return HttpResponseRedirect(reverse('manager:editManagement', args=[id]))
    elif slug == "annoucement":
        Announcement.objects.get(id =id ).image.delete(save = True)
        messages.success(request, "Successfully Cleared Image")
        return HttpResponseRedirect(reverse('manager:editAnnoucement', args=[id]))
def search(request, slug):
    name = request.GET['search']
    if slug == 'branch':
        report = Branch.objects.filter(BranchName__icontains = name)
        type = "branch"
    elif slug == 'surveyor':
        report = Surveryor.objects.filter(name__icontains = name)
        type = "surveyor"
    else:
        report = Agent.objects.filter(name__icontains = name)
        type = "agent"


    dist ={
        'report':report,
        'type':type
       
    }
    return render(request, "manager/search.html", dist)